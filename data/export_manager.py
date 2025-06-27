"""
Advanced Export Manager for Jewelry Database
Supports multiple export formats with advanced filtering and customization options.
"""

import csv
import json
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
import os
import gzip
import zipfile
import logging
from pathlib import Path
from typing import List, Dict, Optional, Any, Union, IO
from datetime import datetime
from enum import Enum
from dataclasses import dataclass, field

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import JewelryListing, JewelryImage, JewelrySpecification
from .query_builder import QueryFilters, QueryBuilder
from .database_manager import DatabaseManager

logger = logging.getLogger(__name__)


class ExportFormat(str, Enum):
    """Supported export formats"""
    CSV = "csv"
    JSON = "json"
    JSONL = "jsonl"  # JSON Lines
    XML = "xml"
    EXCEL = "xlsx"
    PARQUET = "parquet"
    SQLITE = "sqlite"


class CompressionType(str, Enum):
    """Supported compression types"""
    NONE = "none"
    GZIP = "gzip"
    ZIP = "zip"


@dataclass
class ExportOptions:
    """Configuration options for exports"""
    
    # Basic options
    format: ExportFormat = ExportFormat.CSV
    output_path: Optional[str] = None
    compression: CompressionType = CompressionType.NONE
    
    # Content options
    include_metadata: bool = False
    include_raw_data: bool = False
    include_images: bool = False
    include_specifications: bool = False
    flatten_json_fields: bool = True
    
    # Quality filters
    min_quality_score: float = 0.0
    validated_only: bool = False
    
    # Size limits
    max_records: Optional[int] = None
    batch_size: int = 10000
    
    # CSV specific options
    csv_delimiter: str = ","
    csv_quoting: int = csv.QUOTE_MINIMAL
    csv_encoding: str = "utf-8"
    
    # JSON specific options
    json_indent: Optional[int] = 2
    json_ensure_ascii: bool = False
    
    # Excel specific options
    excel_sheet_name: str = "Jewelry Listings"
    excel_freeze_panes: bool = True
    excel_auto_filter: bool = True
    excel_styling: bool = True
    
    # XML specific options
    xml_root_element: str = "jewelry_listings"
    xml_item_element: str = "listing"
    xml_pretty_print: bool = True
    
    # Field selection
    included_fields: Optional[List[str]] = None
    excluded_fields: Optional[List[str]] = None
    
    # Custom field mappings
    field_mappings: Dict[str, str] = field(default_factory=dict)
    
    # Callbacks
    record_transformer: Optional[callable] = None


class ExportManager:
    """
    Advanced export manager with support for multiple formats and advanced options
    """
    
    def __init__(self, database_manager: DatabaseManager):
        """
        Initialize export manager
        
        Args:
            database_manager: DatabaseManager instance
        """
        self.db_manager = database_manager
        self._export_stats = {
            'exports_completed': 0,
            'total_records_exported': 0,
            'total_files_created': 0,
            'last_export_time': None
        }
    
    def export_listings(self, 
                       filters: Optional[QueryFilters] = None,
                       options: Optional[ExportOptions] = None) -> Dict[str, Any]:
        """
        Export jewelry listings with advanced filtering and format options
        
        Args:
            filters: Query filters to apply
            options: Export configuration options
            
        Returns:
            Dict with export results and statistics
        """
        # Set defaults
        filters = filters or QueryFilters()
        options = options or ExportOptions()
        
        # Validate options
        self._validate_export_options(options)
        
        # Generate output path if not provided
        if not options.output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            options.output_path = f"jewelry_export_{timestamp}.{options.format.value}"
        
        # Ensure output directory exists
        output_path = Path(options.output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            logger.info(f"Starting export to {options.output_path} in {options.format.value} format")
            
            # Apply quality filters to main filters
            if options.min_quality_score > 0:
                filters.min_quality_score = options.min_quality_score
            if options.validated_only:
                filters.is_validated = True
            
            # Build query
            query_builder = QueryBuilder().apply_filters(filters)
            
            # Get total count for progress tracking
            count_query, count_params = query_builder.build_count_query()
            with self.db_manager.connection_pool.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(count_query, count_params)
                total_records = cursor.fetchone()[0]
            
            if total_records == 0:
                return {
                    'success': False,
                    'error': 'No records found matching the specified filters',
                    'records_exported': 0
                }
            
            # Apply record limit
            if options.max_records:
                total_records = min(total_records, options.max_records)
                query_builder.limit(options.max_records)
            
            # Perform export based on format
            if options.format == ExportFormat.CSV:
                result = self._export_to_csv(query_builder, options, total_records)
            elif options.format == ExportFormat.JSON:
                result = self._export_to_json(query_builder, options, total_records)
            elif options.format == ExportFormat.JSONL:
                result = self._export_to_jsonl(query_builder, options, total_records)
            elif options.format == ExportFormat.XML:
                result = self._export_to_xml(query_builder, options, total_records)
            elif options.format == ExportFormat.EXCEL:
                result = self._export_to_excel(query_builder, options, total_records)
            elif options.format == ExportFormat.PARQUET:
                result = self._export_to_parquet(query_builder, options, total_records)
            elif options.format == ExportFormat.SQLITE:
                result = self._export_to_sqlite(query_builder, options, total_records)
            else:
                raise ValueError(f"Unsupported export format: {options.format}")
            
            # Apply compression if requested
            if options.compression != CompressionType.NONE:
                result = self._apply_compression(result, options)
            
            # Update statistics
            self._update_export_stats(result)
            
            # Log export record
            self._log_export(options, result)
            
            logger.info(f"Export completed: {result['records_exported']} records exported to {result['output_path']}")
            
            return result
            
        except Exception as e:
            logger.error(f"Export failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'records_exported': 0
            }
    
    def _validate_export_options(self, options: ExportOptions):
        """Validate export options and dependencies"""
        
        # Check format-specific dependencies
        if options.format == ExportFormat.EXCEL and not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        if options.format == ExportFormat.PARQUET and not PANDAS_AVAILABLE:
            raise ValueError("pandas is required for Parquet export. Install with: pip install pandas pyarrow")
        
        # Validate field selections
        if options.included_fields and options.excluded_fields:
            raise ValueError("Cannot specify both included_fields and excluded_fields")
        
        # Validate batch size
        if options.batch_size <= 0:
            raise ValueError("batch_size must be positive")
        
        # Validate quality score
        if not 0 <= options.min_quality_score <= 1:
            raise ValueError("min_quality_score must be between 0 and 1")
    
    def _get_data_in_batches(self, query_builder: QueryBuilder, options: ExportOptions, total_records: int):
        """Generator that yields data in batches"""
        
        offset = 0
        batch_size = options.batch_size
        
        while offset < total_records:
            # Create a copy of the query builder for this batch
            batch_builder = QueryBuilder()
            batch_builder._select_fields = query_builder._select_fields.copy()
            batch_builder._from_table = query_builder._from_table
            batch_builder._table_alias = query_builder._table_alias
            batch_builder._joins = query_builder._joins.copy()
            batch_builder._where_conditions = query_builder._where_conditions.copy()
            batch_builder._sort_options = query_builder._sort_options.copy()
            
            # Set batch pagination
            batch_builder.limit(batch_size).offset(offset)
            
            # Execute query
            query, params = batch_builder.build()
            
            with self.db_manager.connection_pool.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, params)
                
                batch_records = []
                for row in cursor.fetchall():
                    record = dict(row)
                    
                    # Apply field selection
                    record = self._apply_field_selection(record, options)
                    
                    # Apply field mappings
                    record = self._apply_field_mappings(record, options)
                    
                    # Apply record transformer
                    if options.record_transformer:
                        record = options.record_transformer(record)
                    
                    # Process JSON fields
                    if options.flatten_json_fields:
                        record = self._flatten_json_fields(record)
                    
                    batch_records.append(record)
                
                yield batch_records
            
            offset += batch_size
            
            # Respect max records limit
            if options.max_records and offset >= options.max_records:
                break
    
    def _apply_field_selection(self, record: Dict[str, Any], options: ExportOptions) -> Dict[str, Any]:
        """Apply field inclusion/exclusion rules"""
        
        if options.included_fields:
            return {k: v for k, v in record.items() if k in options.included_fields}
        
        if options.excluded_fields:
            return {k: v for k, v in record.items() if k not in options.excluded_fields}
        
        # Default exclusions based on options
        excluded = set()
        if not options.include_metadata:
            excluded.update(['metadata', 'extraction_metadata', 'processing_flags'])
        if not options.include_raw_data:
            excluded.update(['raw_data'])
        
        return {k: v for k, v in record.items() if k not in excluded}
    
    def _apply_field_mappings(self, record: Dict[str, Any], options: ExportOptions) -> Dict[str, Any]:
        """Apply custom field name mappings"""
        
        if not options.field_mappings:
            return record
        
        mapped_record = {}
        for key, value in record.items():
            new_key = options.field_mappings.get(key, key)
            mapped_record[new_key] = value
        
        return mapped_record
    
    def _flatten_json_fields(self, record: Dict[str, Any]) -> Dict[str, Any]:
        """Flatten JSON string fields for better export compatibility"""
        
        json_fields = ['materials', 'features', 'tags', 'accent_stones', 'keywords', 'custom_categories']
        
        for field in json_fields:
            if field in record and record[field]:
                try:
                    if isinstance(record[field], str):
                        parsed = json.loads(record[field])
                        if isinstance(parsed, list):
                            record[field] = ', '.join(str(item) for item in parsed)
                        elif isinstance(parsed, dict):
                            record[field] = json.dumps(parsed, ensure_ascii=False)
                except (json.JSONDecodeError, TypeError):
                    # Keep original value if parsing fails
                    pass
        
        return record
    
    def _export_to_csv(self, query_builder: QueryBuilder, options: ExportOptions, total_records: int) -> Dict[str, Any]:
        """Export data to CSV format"""
        
        records_exported = 0
        
        with open(options.output_path, 'w', newline='', encoding=options.csv_encoding) as csvfile:
            writer = None
            
            for batch_records in self._get_data_in_batches(query_builder, options, total_records):
                if not batch_records:
                    continue
                
                # Initialize writer with first batch
                if writer is None:
                    fieldnames = batch_records[0].keys()
                    writer = csv.DictWriter(
                        csvfile,
                        fieldnames=fieldnames,
                        delimiter=options.csv_delimiter,
                        quoting=options.csv_quoting
                    )
                    writer.writeheader()
                
                # Write batch
                for record in batch_records:
                    # Convert None values to empty strings for CSV
                    csv_record = {k: (v if v is not None else '') for k, v in record.items()}
                    writer.writerow(csv_record)
                    records_exported += 1
        
        return {
            'success': True,
            'records_exported': records_exported,
            'output_path': options.output_path,
            'file_size_mb': Path(options.output_path).stat().st_size / (1024 * 1024)
        }
    
    def _export_to_json(self, query_builder: QueryBuilder, options: ExportOptions, total_records: int) -> Dict[str, Any]:
        """Export data to JSON format"""
        
        all_records = []
        records_exported = 0
        
        for batch_records in self._get_data_in_batches(query_builder, options, total_records):
            all_records.extend(batch_records)
            records_exported += len(batch_records)
        
        # Write JSON file
        with open(options.output_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(
                all_records,
                jsonfile,
                indent=options.json_indent,
                ensure_ascii=options.json_ensure_ascii,
                default=str  # Handle datetime objects
            )
        
        return {
            'success': True,
            'records_exported': records_exported,
            'output_path': options.output_path,
            'file_size_mb': Path(options.output_path).stat().st_size / (1024 * 1024)
        }
    
    def _export_to_jsonl(self, query_builder: QueryBuilder, options: ExportOptions, total_records: int) -> Dict[str, Any]:
        """Export data to JSON Lines format"""
        
        records_exported = 0
        
        with open(options.output_path, 'w', encoding='utf-8') as jsonlfile:
            for batch_records in self._get_data_in_batches(query_builder, options, total_records):
                for record in batch_records:
                    json.dump(record, jsonlfile, ensure_ascii=options.json_ensure_ascii, default=str)
                    jsonlfile.write('\n')
                    records_exported += 1
        
        return {
            'success': True,
            'records_exported': records_exported,
            'output_path': options.output_path,
            'file_size_mb': Path(options.output_path).stat().st_size / (1024 * 1024)
        }
    
    def _export_to_xml(self, query_builder: QueryBuilder, options: ExportOptions, total_records: int) -> Dict[str, Any]:
        """Export data to XML format"""
        
        root = ET.Element(options.xml_root_element)
        records_exported = 0
        
        for batch_records in self._get_data_in_batches(query_builder, options, total_records):
            for record in batch_records:
                item_element = ET.SubElement(root, options.xml_item_element)
                
                for key, value in record.items():
                    field_element = ET.SubElement(item_element, key)
                    field_element.text = str(value) if value is not None else ''
                
                records_exported += 1
        
        # Create XML string
        xml_string = ET.tostring(root, encoding='utf-8', xml_declaration=True)
        
        # Pretty print if requested
        if options.xml_pretty_print:
            dom = minidom.parseString(xml_string)
            xml_string = dom.toprettyxml(indent="  ", encoding='utf-8')
        
        # Write to file
        with open(options.output_path, 'wb') as xmlfile:
            xmlfile.write(xml_string)
        
        return {
            'success': True,
            'records_exported': records_exported,
            'output_path': options.output_path,
            'file_size_mb': Path(options.output_path).stat().st_size / (1024 * 1024)
        }
    
    def _export_to_excel(self, query_builder: QueryBuilder, options: ExportOptions, total_records: int) -> Dict[str, Any]:
        """Export data to Excel format"""
        
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl is required for Excel export")
        
        all_records = []
        records_exported = 0
        
        # Collect all records
        for batch_records in self._get_data_in_batches(query_builder, options, total_records):
            all_records.extend(batch_records)
            records_exported += len(batch_records)
        
        if not all_records:
            raise ValueError("No records to export")
        
        # Create DataFrame
        if PANDAS_AVAILABLE:
            df = pd.DataFrame(all_records)
            
            # Write to Excel with pandas
            with pd.ExcelWriter(options.output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=options.excel_sheet_name, index=False)
                
                # Apply Excel formatting if requested
                if options.excel_styling:
                    self._apply_excel_styling(writer, options.excel_sheet_name, df)
        
        else:
            # Use openpyxl directly
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = options.excel_sheet_name
            
            # Write headers
            headers = list(all_records[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row_idx, record in enumerate(all_records, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=record.get(header))
            
            # Apply formatting
            if options.excel_styling:
                self._apply_excel_styling_direct(ws)
            
            # Apply filters and freeze panes
            if options.excel_auto_filter:
                ws.auto_filter.ref = ws.dimensions
            
            if options.excel_freeze_panes:
                ws.freeze_panes = 'A2'
            
            wb.save(options.output_path)
        
        return {
            'success': True,
            'records_exported': records_exported,
            'output_path': options.output_path,
            'file_size_mb': Path(options.output_path).stat().st_size / (1024 * 1024)
        }
    
    def _apply_excel_styling(self, writer, sheet_name: str, df):
        """Apply styling to Excel worksheet using pandas ExcelWriter"""
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_excel_styling_direct(self, worksheet):
        """Apply styling directly to openpyxl worksheet"""
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply to first row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _export_to_parquet(self, query_builder: QueryBuilder, options: ExportOptions, total_records: int) -> Dict[str, Any]:
        """Export data to Parquet format"""
        
        if not PANDAS_AVAILABLE:
            raise ValueError("pandas is required for Parquet export")
        
        all_records = []
        records_exported = 0
        
        for batch_records in self._get_data_in_batches(query_builder, options, total_records):
            all_records.extend(batch_records)
            records_exported += len(batch_records)
        
        # Create DataFrame and save as Parquet
        df = pd.DataFrame(all_records)
        df.to_parquet(options.output_path, index=False)
        
        return {
            'success': True,
            'records_exported': records_exported,
            'output_path': options.output_path,
            'file_size_mb': Path(options.output_path).stat().st_size / (1024 * 1024)
        }
    
    def _export_to_sqlite(self, query_builder: QueryBuilder, options: ExportOptions, total_records: int) -> Dict[str, Any]:
        """Export data to SQLite database"""
        
        import sqlite3
        
        records_exported = 0
        
        # Create new SQLite database
        conn = sqlite3.connect(options.output_path)
        cursor = conn.cursor()
        
        # Create table schema based on first batch
        table_created = False
        
        for batch_records in self._get_data_in_batches(query_builder, options, total_records):
            if not batch_records:
                continue
            
            # Create table on first batch
            if not table_created:
                self._create_sqlite_table(cursor, batch_records[0], "exported_listings")
                table_created = True
            
            # Insert batch
            for record in batch_records:
                placeholders = ', '.join(['?' for _ in record])
                columns = ', '.join(record.keys())
                
                cursor.execute(
                    f"INSERT INTO exported_listings ({columns}) VALUES ({placeholders})",
                    list(record.values())
                )
                records_exported += 1
        
        # Create indexes for better performance
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_listing_id ON exported_listings(listing_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_category ON exported_listings(category)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_price ON exported_listings(price)")
        
        conn.commit()
        conn.close()
        
        return {
            'success': True,
            'records_exported': records_exported,
            'output_path': options.output_path,
            'file_size_mb': Path(options.output_path).stat().st_size / (1024 * 1024)
        }
    
    def _create_sqlite_table(self, cursor, sample_record: Dict[str, Any], table_name: str):
        """Create SQLite table based on sample record"""
        
        # Map Python types to SQLite types
        type_mapping = {
            str: 'TEXT',
            int: 'INTEGER',
            float: 'REAL',
            bool: 'BOOLEAN',
            type(None): 'TEXT'
        }
        
        columns = []
        for key, value in sample_record.items():
            sql_type = type_mapping.get(type(value), 'TEXT')
            columns.append(f"{key} {sql_type}")
        
        create_sql = f"CREATE TABLE {table_name} ({', '.join(columns)})"
        cursor.execute(create_sql)
    
    def _apply_compression(self, result: Dict[str, Any], options: ExportOptions) -> Dict[str, Any]:
        """Apply compression to the export file"""
        
        original_path = result['output_path']
        
        if options.compression == CompressionType.GZIP:
            compressed_path = original_path + '.gz'
            
            with open(original_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    f_out.writelines(f_in)
            
            # Remove original file
            os.remove(original_path)
            result['output_path'] = compressed_path
            result['compression'] = 'gzip'
        
        elif options.compression == CompressionType.ZIP:
            compressed_path = original_path + '.zip'
            
            with zipfile.ZipFile(compressed_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(original_path, Path(original_path).name)
            
            # Remove original file
            os.remove(original_path)
            result['output_path'] = compressed_path
            result['compression'] = 'zip'
        
        # Update file size
        result['file_size_mb'] = Path(result['output_path']).stat().st_size / (1024 * 1024)
        
        return result
    
    def _update_export_stats(self, result: Dict[str, Any]):
        """Update export statistics"""
        
        if result.get('success'):
            self._export_stats['exports_completed'] += 1
            self._export_stats['total_records_exported'] += result.get('records_exported', 0)
            self._export_stats['total_files_created'] += 1
            self._export_stats['last_export_time'] = datetime.now()
    
    def _log_export(self, options: ExportOptions, result: Dict[str, Any]):
        """Log export operation to database"""
        
        try:
            with self.db_manager.connection_pool.get_connection() as conn:
                cursor = conn.cursor()
                
                cursor.execute("""
                    INSERT INTO export_history (
                        export_id, export_type, file_path, record_count, file_size_mb,
                        status, created_at, created_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                    options.format.value,
                    result.get('output_path'),
                    result.get('records_exported', 0),
                    result.get('file_size_mb', 0.0),
                    'completed' if result.get('success') else 'failed',
                    datetime.now().isoformat(),
                    'system'
                ))
                
                conn.commit()
        
        except Exception as e:
            logger.warning(f"Failed to log export operation: {e}")
    
    def get_export_stats(self) -> Dict[str, Any]:
        """Get export statistics"""
        return self._export_stats.copy()
    
    def get_export_history(self, limit: int = 10) -> List[Dict[str, Any]]:
        """Get recent export history"""
        
        try:
            with self.db_manager.connection_pool.get_connection() as conn:
                cursor = conn.cursor()
                
                cursor.execute("""
                    SELECT * FROM export_history
                    ORDER BY created_at DESC
                    LIMIT ?
                """, (limit,))
                
                return [dict(row) for row in cursor.fetchall()]
        
        except Exception as e:
            logger.error(f"Failed to get export history: {e}")
            return []


# Convenience functions for common export scenarios
def export_high_quality_listings(db_manager: DatabaseManager, 
                                output_path: str = None,
                                format: ExportFormat = ExportFormat.CSV) -> Dict[str, Any]:
    """Export high-quality listings"""
    
    exporter = ExportManager(db_manager)
    
    filters = QueryFilters(
        min_quality_score=0.8,
        is_validated=True,
        has_images=True
    )
    
    options = ExportOptions(
        format=format,
        output_path=output_path,
        include_metadata=False,
        validated_only=True
    )
    
    return exporter.export_listings(filters, options)


def export_category_summary(db_manager: DatabaseManager,
                           category: str,
                           output_path: str = None) -> Dict[str, Any]:
    """Export summary for a specific category"""
    
    exporter = ExportManager(db_manager)
    
    filters = QueryFilters(
        category=category,
        is_validated=True
    )
    
    # Define summary fields
    summary_fields = [
        'listing_id', 'title', 'price', 'currency', 'brand', 'seller_name',
        'condition', 'main_stone', 'size', 'image_count', 'watchers', 'views',
        'data_completeness_score', 'scraped_at', 'url'
    ]
    
    options = ExportOptions(
        format=ExportFormat.EXCEL,
        output_path=output_path or f"{category}_summary.xlsx",
        included_fields=summary_fields,
        excel_styling=True,
        excel_auto_filter=True
    )
    
    return exporter.export_listings(filters, options)